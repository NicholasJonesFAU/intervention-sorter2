"""
excel_utils.py — openpyxl formatting helpers.
All colors use full 8-char ARGB format (FF prefix) required by Excel.
"""

from typing import List, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils.config import STYLE, COLUMN_WIDTH_OVERRIDES


def _argb(hex6: str) -> str:
    """Convert a 6-char RGB hex to 8-char ARGB by prepending FF."""
    hex6 = hex6.lstrip("#")
    return f"FF{hex6.upper()}" if len(hex6) == 6 else hex6.upper()


def make_header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=_argb(hex_color))


def make_alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_argb(STYLE.alt_row_fill_color))


def make_default_fill() -> PatternFill:
    return PatternFill("solid", fgColor=_argb(STYLE.default_fill_color))


def make_header_font(hex_color: str = None) -> Font:
    color = _argb(hex_color or STYLE.header_font_color)
    return Font(
        name=STYLE.header_font_name,
        size=STYLE.header_font_size,
        bold=True,
        color=color,
    )


def make_body_font() -> Font:
    return Font(name=STYLE.body_font_name, size=STYLE.body_font_size)


def make_wrap_alignment(horizontal: str = "left") -> Alignment:
    return Alignment(wrap_text=True, vertical="top", horizontal=horizontal)


def apply_data_tab_formatting(
    ws: Worksheet,
    columns: List[str],
    header_hex: str = None,
) -> None:
    header_fill = make_header_fill(header_hex or STYLE.header_fill_color)
    header_font = make_header_font()
    alt_fill = make_alt_fill()
    default_fill = make_default_fill()
    body_font = make_body_font()
    wrap_align = make_wrap_alignment()

    max_col = ws.max_column
    max_row = ws.max_row

    # Header row
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = make_wrap_alignment("center")

    # Data rows
    for row_idx in range(2, max_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else default_fill
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.font = body_font
            cell.alignment = wrap_align
        ws.row_dimensions[row_idx].height = STYLE.row_height

    # Freeze header
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Autofilter
    if max_col > 0:
        ws.auto_filter.ref = ws.dimensions

    # Column widths
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in COLUMN_WIDTH_OVERRIDES:
            ws.column_dimensions[col_letter].width = COLUMN_WIDTH_OVERRIDES[col_name]
        else:
            max_len = len(col_name)
            for row_idx in range(2, min(max_row + 1, 200)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    lines = str(val).split("\n")
                    max_len = max(max_len, max(len(l) for l in lines))
            ws.column_dimensions[col_letter].width = max(
                STYLE.min_col_width, min(max_len + 2, STYLE.max_col_width)
            )


def apply_summary_formatting(ws: Worksheet) -> None:
    header_fill = make_header_fill(STYLE.summary_accent_color)
    header_font = Font(
        name=STYLE.header_font_name,
        size=12,
        bold=True,
        color=_argb(STYLE.header_font_color),
    )
    body_font = Font(name=STYLE.body_font_name, size=10)
    label_font = Font(name=STYLE.body_font_name, size=10, bold=True)

    for row in ws.iter_rows():
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.font = label_font

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("◆"):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = make_wrap_alignment("center")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 35


def apply_qa_formatting(ws: Worksheet) -> None:
    apply_data_tab_formatting(
        ws,
        columns=["Category", "Student ID", "Detail", "Source File", "Timestamp"],
        header_hex=STYLE.qa_header_color,
    )


def apply_manifest_formatting(ws: Worksheet) -> None:
    apply_summary_formatting(ws)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).startswith("◆"):
                cell.fill = make_header_fill(STYLE.manifest_header_color)
