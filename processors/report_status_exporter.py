"""
report_status_exporter.py — Builds the faculty completion Excel workbook.

Tabs:
    Workbook_Index   — clickable navigation + workbook metadata
    Overview         — KPI cards, overall stats, donut chart, college table
    By_College       — college table + horizontal bar chart
    By_Department    — department table + horizontal bar chart
    By_Professor     — full professor detail table
    Faculty_Download — one row per faculty: First, Last, Email, College, Dept
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

from processors.report_status_processor import ReportStatusProcessor
from utils.excel_utils import _argb

logger = logging.getLogger("intervention_sorter")

# Chart color palette
COLOR_SUBMITTED = "#2E86AB"
COLOR_NOT_SUBMITTED = "#E84855"
COLOR_BG = "#F4F6FB"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_BLUE = "DCE6F1"
COLOR_LIGHT_GRAY = "ECEFF1"
COLOR_GREEN = "2E7D32"
COLOR_AMBER = "F57F17"
COLOR_RED = "C62828"

HEADER_COLOR = "1F3864"
COLLEGE_HEADER = "2F5496"
DEPT_HEADER = "375623"
PROFESSOR_HEADER = "843C0C"
DOWNLOAD_HEADER = "4A235A"
INDEX_HEADER = "1F3864"

THIN_WHITE_BORDER = Border(
    left=Side(style="thin", color=_argb("FFFFFF")),
    right=Side(style="thin", color=_argb("FFFFFF")),
    top=Side(style="thin", color=_argb("FFFFFF")),
    bottom=Side(style="thin", color=_argb("FFFFFF")),
)


class ReportStatusExporter:

    def export(
        self,
        processor: ReportStatusProcessor,
        output_path: Path,
        source_filename: str,
    ) -> None:
        logger.info("ReportStatusExporter: Building workbook → '%s'", output_path.name)

        wb = Workbook()
        wb.remove(wb.active)

        overall = processor.overall_stats()
        df_college = processor.by_college()
        df_dept = processor.by_department()
        df_prof = processor.by_professor()
        df_dl = processor.faculty_download()

        self._write_index(
            wb=wb,
            source_filename=source_filename,
            overall=overall,
            sheet_rows=[
                (
                    "Overview",
                    "Executive completion summary, KPI cards, and college rollup.",
                ),
                ("By_College", "Completion counts and rates grouped by college."),
                ("By_Department", "Completion counts and rates grouped by department."),
                ("By_Professor", "Faculty-level completion detail."),
                ("Faculty_Download", "Clean faculty contact list for outreach/export."),
            ],
        )

        self._write_overview(wb, overall, df_college, source_filename)
        self._write_college_tab(wb, df_college)
        self._write_department_tab(wb, df_dept)
        self._write_professor_tab(wb, df_prof)
        self._write_faculty_download(wb, df_dl)

        try:
            wb.save(output_path)
            logger.info("ReportStatusExporter: Saved → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Cannot save — file may be open in Excel.\nPath: {output_path}\n{exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Workbook Index
    # ------------------------------------------------------------------

    def _write_index(
        self,
        wb: Workbook,
        source_filename: str,
        overall: Dict[str, Any],
        sheet_rows: List[Tuple[str, str]],
    ) -> None:
        ws = wb.create_sheet("Workbook_Index")

        ws.merge_cells("A1:E1")
        c = ws["A1"]
        c.value = "Faculty Completion Workbook"
        c.font = Font(name="Calibri", size=15, bold=True, color=_argb(COLOR_WHITE))
        c.fill = PatternFill("solid", fgColor=_argb(INDEX_HEADER))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        metadata = [
            ("Generated", datetime.now().strftime("%m/%d/%Y %I:%M %p")),
            ("Source File", source_filename),
            ("Total Sections", overall.get("total_sections", 0)),
            ("Submitted", overall.get("submitted", 0)),
            ("Not Submitted", overall.get("not_submitted", 0)),
            ("Overall Completion", f"{overall.get('completion_pct', 0)}%"),
        ]

        r = 3
        for label, value in metadata:
            ws.cell(r, 1, label).font = Font(
                name="Calibri", size=10, bold=True, color=_argb(HEADER_COLOR)
            )
            ws.cell(r, 2, value).font = Font(name="Calibri", size=10)
            r += 1

        r += 1
        headers = ["Sheet", "Description", "Rows / Scope"]
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(r, idx, header)
            cell.font = Font(
                name="Calibri", size=10, bold=True, color=_argb(COLOR_WHITE)
            )
            cell.fill = PatternFill("solid", fgColor=_argb(INDEX_HEADER))
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_WHITE_BORDER

        for sheet_name, desc in sheet_rows:
            r += 1
            link_cell = ws.cell(r, 1, sheet_name)
            link_cell.hyperlink = f"#'{sheet_name}'!A1"
            link_cell.style = "Hyperlink"
            ws.cell(r, 2, desc)

            if sheet_name == "Overview":
                scope = "Executive summary"
            elif sheet_name == "By_College":
                scope = "College rollup"
            elif sheet_name == "By_Department":
                scope = "Department rollup"
            elif sheet_name == "By_Professor":
                scope = "Faculty detail"
            else:
                scope = "Download-ready list"

            ws.cell(r, 3, scope)

            fill = PatternFill(
                "solid", fgColor=_argb(COLOR_LIGHT_BLUE if r % 2 == 0 else COLOR_WHITE)
            )
            for col in range(1, 4):
                ws.cell(r, col).fill = fill
                ws.cell(r, col).alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 64
        ws.column_dimensions["C"].width = 24
        ws.freeze_panes = "A11"

    # ------------------------------------------------------------------
    # Overview tab
    # ------------------------------------------------------------------

    def _write_overview(
        self,
        wb: Workbook,
        overall: Dict[str, Any],
        df_college: pd.DataFrame,
        source_filename: str,
    ) -> None:
        ws = wb.create_sheet("Overview")

        # Header banner
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = "Faculty Progress Report — Completion Overview"
        c.font = Font(name="Calibri", size=15, bold=True, color=_argb(COLOR_WHITE))
        c.fill = PatternFill("solid", fgColor=_argb(HEADER_COLOR))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Metadata
        ws.cell(2, 1, "Source File").font = Font(
            name="Calibri", size=10, bold=True, color=_argb(HEADER_COLOR)
        )
        ws.cell(2, 2, source_filename)
        ws.cell(3, 1, "Generated").font = Font(
            name="Calibri", size=10, bold=True, color=_argb(HEADER_COLOR)
        )
        ws.cell(3, 2, datetime.now().strftime("%m/%d/%Y %I:%M %p"))

        # KPI cards
        self._write_kpi_cards(ws, overall, start_row=5)

        # Overall donut chart
        chart_img = self._make_donut_chart(
            submitted=overall["submitted"],
            not_submitted=overall["not_submitted"],
            pct=overall["completion_pct"],
            title="Overall Submission Rate",
        )
        ws.add_image(chart_img, "E3")

        # College summary table below
        start_row = 17
        ws.merge_cells(
            start_row=start_row, start_column=1, end_row=start_row, end_column=5
        )
        section = ws.cell(start_row, 1)
        section.value = "Completion by College"
        section.font = Font(
            name="Calibri", size=11, bold=True, color=_argb(COLOR_WHITE)
        )
        section.fill = PatternFill("solid", fgColor=_argb(COLLEGE_HEADER))
        section.alignment = Alignment(horizontal="center")

        headers = [
            "College",
            "Total Sections",
            "Submitted",
            "Not Submitted",
            "Completion %",
        ]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(start_row + 1, c_idx, h)
            cell.font = Font(
                name="Calibri", size=10, bold=True, color=_argb(COLOR_WHITE)
            )
            cell.fill = PatternFill("solid", fgColor=_argb(COLLEGE_HEADER))
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_WHITE_BORDER

        for r_idx, row in df_college.iterrows():
            excel_row = start_row + 2 + r_idx
            fill = PatternFill(
                "solid",
                fgColor=_argb(COLOR_LIGHT_BLUE if r_idx % 2 == 0 else COLOR_WHITE),
            )
            values = [
                row["College"],
                row["Total Sections"],
                row["Submitted"],
                row["Not Submitted"],
                f"{row['Completion %']}%",
            ]
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(excel_row, c_idx, val)
                cell.fill = fill
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
                if c_idx == 5:
                    self._apply_completion_fill(cell, row["Completion %"])

        for c_idx, w in enumerate([26, 16, 13, 16, 14], 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.column_dimensions["F"].width = 4
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 18

        ws.freeze_panes = "A17"

    def _write_kpi_cards(
        self,
        ws,
        overall: Dict[str, Any],
        start_row: int,
    ) -> None:
        cards = [
            ("Total Sections", overall.get("total_sections", 0), HEADER_COLOR),
            ("Submitted", overall.get("submitted", 0), COLOR_GREEN),
            ("Not Submitted", overall.get("not_submitted", 0), COLOR_RED),
            (
                "Completion",
                f"{overall.get('completion_pct', 0)}%",
                self._completion_color(overall.get("completion_pct", 0)),
            ),
        ]

        col = 1
        for label, value, color in cards:
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row,
                end_column=col + 1,
            )
            ws.merge_cells(
                start_row=start_row + 1,
                start_column=col,
                end_row=start_row + 2,
                end_column=col + 1,
            )

            label_cell = ws.cell(start_row, col)
            label_cell.value = label
            label_cell.font = Font(
                name="Calibri", size=10, bold=True, color=_argb(COLOR_WHITE)
            )
            label_cell.fill = PatternFill("solid", fgColor=_argb(color))
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            label_cell.border = THIN_WHITE_BORDER

            value_cell = ws.cell(start_row + 1, col)
            value_cell.value = value
            value_cell.font = Font(
                name="Calibri", size=16, bold=True, color=_argb(HEADER_COLOR)
            )
            value_cell.fill = PatternFill("solid", fgColor=_argb(COLOR_LIGHT_GRAY))
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            value_cell.border = THIN_WHITE_BORDER

            col += 2

    # ------------------------------------------------------------------
    # By College tab
    # ------------------------------------------------------------------

    def _write_college_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_College")
        self._write_completion_table(ws, df, COLLEGE_HEADER, "Completion by College")

        chart_img = self._make_bar_chart(
            df,
            label_col="College",
            title="Submission Rate by College",
            figsize=(7, max(3, len(df) * 0.35 + 1)),
        )
        ws.add_image(chart_img, f"A{len(df) + 6}")

    # ------------------------------------------------------------------
    # By Department tab
    # ------------------------------------------------------------------

    def _write_department_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_Department")
        df_display = df.copy()
        self._write_completion_table(
            ws,
            df_display,
            DEPT_HEADER,
            "Completion by Department",
            extra_cols=["College", "Department"],
        )

        chart_img = self._make_bar_chart(
            df,
            label_col="Department",
            title="Submission Rate by Department",
            figsize=(8, max(4, len(df) * 0.3 + 1)),
        )
        ws.add_image(chart_img, f"A{len(df) + 6}")

    # ------------------------------------------------------------------
    # By Professor tab
    # ------------------------------------------------------------------

    def _write_professor_tab(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("By_Professor")
        cols = [
            "Last Name",
            "First Name",
            "Email",
            "College",
            "Department",
            "Total Sections",
            "Submitted",
            "Not Submitted",
            "Completion %",
        ]
        df_out = df[cols].copy()
        df_out["Completion %"] = df_out["Completion %"].apply(lambda x: f"{x}%")
        self._write_table(ws, df_out, PROFESSOR_HEADER, "Faculty Submission Detail")

    # ------------------------------------------------------------------
    # Faculty Download tab
    # ------------------------------------------------------------------

    def _write_faculty_download(self, wb: Workbook, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Faculty_Download")
        self._write_table(ws, df, DOWNLOAD_HEADER, "Faculty Contact Download")

    # ------------------------------------------------------------------
    # Shared table writer
    # ------------------------------------------------------------------

    def _write_completion_table(
        self,
        ws,
        df: pd.DataFrame,
        header_color: str,
        title: str,
        extra_cols: list = None,
    ) -> None:
        base_cols = ["Total Sections", "Submitted", "Not Submitted", "Completion %"]

        if extra_cols:
            display_cols = extra_cols + base_cols
        else:
            display_cols = [c for c in df.columns if c in base_cols] or list(df.columns)

        df_out = df[[c for c in display_cols if c in df.columns]].copy()
        df_out["Completion %"] = df_out["Completion %"].apply(lambda x: f"{x}%")
        self._write_table(ws, df_out, header_color, title)

    def _write_table(
        self,
        ws,
        df: pd.DataFrame,
        header_color: str,
        title: str,
    ) -> None:
        max_col = max(len(df.columns), 1)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(1, 1, title)
        title_cell.font = Font(
            name="Calibri", size=13, bold=True, color=_argb(COLOR_WHITE)
        )
        title_cell.fill = PatternFill("solid", fgColor=_argb(header_color))
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Headers
        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(
                name="Calibri", size=10, bold=True, color=_argb(COLOR_WHITE)
            )
            cell.fill = PatternFill("solid", fgColor=_argb(header_color))
            cell.alignment = Alignment(
                horizontal="center" if c_idx > 2 else "left", vertical="center"
            )
            cell.border = THIN_WHITE_BORDER

        # Freeze and filter
        ws.freeze_panes = "A3"
        if len(df.columns) > 0:
            ws.auto_filter.ref = (
                f"A2:{get_column_letter(len(df.columns))}{max(ws.max_row, len(df) + 2)}"
            )

        alt_fill = PatternFill("solid", fgColor=_argb(COLOR_LIGHT_BLUE))
        white_fill = PatternFill("solid", fgColor=_argb(COLOR_WHITE))
        body_font = Font(name="Calibri", size=10)

        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            fill = alt_fill if r_idx % 2 == 1 else white_fill

            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, str(val) if pd.notna(val) else "")
                cell.fill = fill
                cell.font = body_font
                cell.alignment = Alignment(
                    horizontal="center" if c_idx > 2 else "left",
                    vertical="top",
                    wrap_text=True,
                )

                col_name = df.columns[c_idx - 1]
                if col_name == "Completion %":
                    numeric = self._extract_pct(val)
                    self._apply_completion_fill(cell, numeric)

        # Auto column widths
        for c_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            for r_idx in range(3, min(ws.max_row + 1, 300)):
                val = ws.cell(r_idx, c_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(
                max(max_len + 3, 12), 50
            )

        # Metadata/footer
        footer_row = ws.max_row + 2
        ws.cell(
            footer_row, 1, f"Generated {datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
        )
        ws.cell(footer_row, 1).font = Font(
            name="Calibri", size=9, italic=True, color=_argb("666666")
        )

    # ------------------------------------------------------------------
    # Styling helpers
    # ------------------------------------------------------------------

    def _completion_color(self, pct: float) -> str:
        try:
            pct = float(pct)
        except Exception:
            pct = 0

        if pct >= 90:
            return COLOR_GREEN
        if pct >= 70:
            return COLOR_AMBER
        return COLOR_RED

    def _apply_completion_fill(self, cell, pct: float) -> None:
        color = self._completion_color(pct)
        cell.fill = PatternFill("solid", fgColor=_argb(color))
        cell.font = Font(name="Calibri", size=10, bold=True, color=_argb(COLOR_WHITE))
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _extract_pct(self, value: Any) -> float:
        try:
            if isinstance(value, str):
                return float(value.replace("%", "").strip())
            return float(value)
        except Exception:
            return 0.0

    # ------------------------------------------------------------------
    # Chart generators
    # ------------------------------------------------------------------

    def _chart_to_image(
        self,
        fig,
        width: int = 500,
        height: int = 315,
    ) -> XLImage:
        """Convert matplotlib figure to openpyxl Image object."""
        buf = io.BytesIO()
        fig.savefig(
            buf,
            format="png",
            dpi=96,
            bbox_inches="tight",
            facecolor=COLOR_BG,
        )
        buf.seek(0)
        plt.close(fig)

        img = XLImage(buf)
        img.width = width
        img.height = height
        return img

    def _make_donut_chart(
        self,
        submitted: int,
        not_submitted: int,
        pct: float,
        title: str,
    ) -> XLImage:
        fig, ax = plt.subplots(figsize=(4.4, 3.3), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)

        sizes = (
            [submitted, not_submitted] if (submitted + not_submitted) > 0 else [1, 0]
        )
        colors = [COLOR_SUBMITTED, COLOR_NOT_SUBMITTED]

        ax.pie(
            sizes,
            colors=colors,
            startangle=90,
            wedgeprops=dict(width=0.5, edgecolor="white", linewidth=2),
        )

        ax.text(
            0,
            0,
            f"{pct}%",
            ha="center",
            va="center",
            fontsize=22,
            fontweight="bold",
            color="#1F3864",
        )

        ax.set_title(
            title,
            fontsize=12,
            fontweight="bold",
            color="#1F3864",
            pad=12,
        )

        legend = [
            mpatches.Patch(color=COLOR_SUBMITTED, label=f"Submitted ({submitted:,})"),
            mpatches.Patch(
                color=COLOR_NOT_SUBMITTED, label=f"Not Submitted ({not_submitted:,})"
            ),
        ]

        ax.legend(
            handles=legend,
            loc="lower center",
            bbox_to_anchor=(0.5, -0.08),
            ncol=2,
            fontsize=9,
            framealpha=0,
        )

        return self._chart_to_image(fig, width=420, height=300)

    def _make_bar_chart(
        self,
        df: pd.DataFrame,
        label_col: str,
        title: str,
        figsize: tuple = (10, 6),
    ) -> XLImage:
        labels = df[label_col].astype(str).tolist()
        submitted = df["Submitted"].tolist()
        not_submitted = df["Not Submitted"].tolist()
        pcts = df["Completion %"].tolist()

        y = np.arange(len(labels))
        height = 0.35

        fig, ax = plt.subplots(figsize=figsize, facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)

        ax.barh(
            y + height / 2,
            submitted,
            height,
            label="Submitted",
            color=COLOR_SUBMITTED,
            edgecolor="white",
        )

        ax.barh(
            y - height / 2,
            not_submitted,
            height,
            label="Not Submitted",
            color=COLOR_NOT_SUBMITTED,
            edgecolor="white",
        )

        for i, pct in enumerate(pcts):
            right_edge = max(submitted[i], not_submitted[i])
            ax.text(
                right_edge + 0.5,
                y[i],
                f"{pct}%",
                va="center",
                ha="left",
                fontsize=8,
                fontweight="bold",
                color="#1F3864",
            )

        ax.set_yticks(y)
        ax.set_yticklabels(labels, fontsize=9)
        ax.set_xlabel("Number of Sections", fontsize=10)
        ax.set_title(title, fontsize=12, fontweight="bold", color="#1F3864", pad=10)
        ax.legend(fontsize=9, framealpha=0.5)

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.invert_yaxis()

        plt.tight_layout()
        return self._chart_to_image(fig)
