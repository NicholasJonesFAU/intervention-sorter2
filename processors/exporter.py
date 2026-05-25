"""
exporter.py — Writes the final output Excel workbook.

Tab order:
  Workbook_Index → Summary → <Group Tabs> → Risk_1_2 → Risk_3_Plus
  → Missing_Contacts (when needed) → QA_Log → Processing_Manifest

All student data tabs use the standardized OUTPUT_COLUMNS schema.
"""

import logging
import platform
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from utils.config import (
    APP_NAME,
    APP_VERSION,
    OUTPUT_COLUMNS,
    SUMMARY_TAB,
    QA_LOG_TAB,
    MANIFEST_TAB,
    UNMATCHED_LOW_TAB,
    UNMATCHED_HIGH_TAB,
    QA_LOG_COLUMNS,
    SUMMARY_LABELS,
    STYLE,
)
from utils.normalization import safe_excel_tab_name
from utils.excel_utils import (
    _argb,
    apply_data_tab_formatting,
    apply_summary_formatting,
    apply_qa_formatting,
    apply_manifest_formatting,
    make_header_fill,
    make_header_font,
    make_body_font,
)
from utils.logging_utils import QALog
from processors.summary_enhancer import SummaryEnhancer

logger = logging.getLogger("intervention_sorter")

INDEX_TAB = "Workbook_Index"
MISSING_CONTACTS_TAB = "Missing_Contacts"

HEADER_DARK = "1F3864"
HEADER_BLUE = "2F5496"
HEADER_GREEN = "375623"
HEADER_ORANGE = "843C0C"
HEADER_PURPLE = "4A235A"
LIGHT_BLUE = "DCE6F1"
LIGHT_FILL = "F4F6FB"
WHITE = "FFFFFF"
BORDER_GRAY = "B7B7B7"


class Exporter:
    """Generates the final output workbook from processed data."""

    def export(
        self,
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        qa_log: QALog,
        metrics: Dict[str, Any],
        output_path: Path,
        source_files: Dict[str, str],
    ) -> None:
        """
        Build and save the output workbook.

        Args:
            group_data:   dict mapping safe_tab_name → DataFrame
            group_order:  ordered list of group tab names, excluding unmatched/QA/manifest
            qa_log:       QALog instance with all QA events
            metrics:      dict of processing metrics for summary/manifest
            output_path:  resolved output .xlsx path
            source_files: dict mapping label → filename for manifest
        """
        logger.info("Exporter: Building output workbook → '%s'", output_path.name)

        wb = Workbook()
        wb.remove(wb.active)

        used_tab_names: List[str] = []
        index_entries: List[Dict[str, Any]] = []

        # 1. Summary tab
        summary_title = self._write_summary(
            wb, metrics, source_files, group_data, group_order, used_tab_names
        )
        index_entries.append(self._index_entry(
            summary_title,
            "Executive summary with source files, key counts, contact coverage, and charts.",
            "Summary",
            "",
        ))

        # 2. Group tabs, in priority/order from the control file or semester setup
        for tab_name in group_order:
            df = group_data.get(tab_name, pd.DataFrame())
            actual_title = self._write_data_tab(wb, tab_name, df, used_tab_names)
            index_entries.append(self._index_entry(
                actual_title,
                "Assigned students for this intervention group.",
                "Group",
                len(df),
            ))

        # 3. Unmatched buckets
        for bucket_name, description in [
            (UNMATCHED_LOW_TAB, "Unmatched students with 1–2 risk courses."),
            (UNMATCHED_HIGH_TAB, "Unmatched students with 3 or more risk courses."),
        ]:
            df = group_data.get(bucket_name, pd.DataFrame())
            actual_title = self._write_data_tab(wb, bucket_name, df, used_tab_names)
            index_entries.append(self._index_entry(
                actual_title,
                description,
                "Unmatched",
                len(df),
            ))

        # 4. Missing Contacts, only when needed
        missing_title, missing_count = self._write_missing_contacts(
            wb, group_data, group_order, used_tab_names
        )
        if missing_title:
            index_entries.append(self._index_entry(
                missing_title,
                "Students without phone or email contact information in the contact report.",
                "QA",
                missing_count,
            ))

        # 5. QA_Log
        qa_title = self._write_qa_log(wb, qa_log, used_tab_names)
        index_entries.append(self._index_entry(
            qa_title,
            "Validation notes, data quality warnings, and processing audit items.",
            "QA",
            len(qa_log.entries()),
        ))

        # 6. Processing_Manifest
        manifest_title = self._write_manifest(wb, metrics, source_files, used_tab_names)
        index_entries.append(self._index_entry(
            manifest_title,
            "Technical run metadata: app version, Python version, platform, inputs, and row counts.",
            "Manifest",
            "",
        ))

        # 7. Workbook index goes first, after the final sheet names are known
        self._write_workbook_index(wb, index_entries, metrics, source_files, used_tab_names)

        try:
            wb.save(output_path)
            logger.info("Exporter: Workbook saved successfully → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Cannot save workbook. The file may be open in Excel.\n"
                f"Path: {output_path}\nError: {exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Workbook index
    # ------------------------------------------------------------------

    def _write_workbook_index(
        self,
        wb: Workbook,
        index_entries: List[Dict[str, Any]],
        metrics: Dict[str, Any],
        source_files: Dict[str, str],
        used_tab_names: List[str],
    ) -> str:
        """Create a clickable table of contents as the first worksheet."""
        title = safe_excel_tab_name(INDEX_TAB, used_tab_names)
        used_tab_names.append(title)
        ws = wb.create_sheet(title=title, index=0)

        # Title banner
        ws.merge_cells("A1:E1")
        cell = ws["A1"]
        cell.value = f"{APP_NAME} — Workbook Index"
        cell.font = Font(name="Calibri", size=15, bold=True, color=_argb(WHITE))
        cell.fill = PatternFill("solid", fgColor=_argb(HEADER_DARK))
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Small KPI strip
        assigned = self._to_int(metrics.get("total_assigned", 0))
        unmatched = self._to_int(metrics.get("total_unmatched", 0))
        distinct = self._to_int(metrics.get("total_distinct_students", 0))
        contact_matches = self._to_int(metrics.get("contact_matches", 0))
        contact_misses = self._to_int(metrics.get("contact_misses", 0))
        contact_pct = self._percent(contact_matches, contact_matches + contact_misses)
        unmatched_pct = self._percent(unmatched, distinct)

        kpis = [
            ("Distinct Students", distinct),
            ("Assigned", assigned),
            ("Unmatched", f"{unmatched} ({unmatched_pct})"),
            ("Contact Coverage", contact_pct),
            ("Generated", metrics.get("processing_timestamp") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        for col_idx, (label, value) in enumerate(kpis, start=1):
            label_cell = ws.cell(row=3, column=col_idx, value=label)
            value_cell = ws.cell(row=4, column=col_idx, value=value)
            label_cell.fill = PatternFill("solid", fgColor=_argb(HEADER_BLUE))
            label_cell.font = Font(name="Calibri", size=10, bold=True, color=_argb(WHITE))
            label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            value_cell.fill = PatternFill("solid", fgColor=_argb(LIGHT_BLUE))
            value_cell.font = Font(name="Calibri", size=11, bold=True)
            value_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 24
        ws.row_dimensions[4].height = 30

        # Navigation table
        start_row = 7
        headers = ["Sheet", "Type", "Rows", "Description", "Open"]
        for col_idx, header in enumerate(headers, start=1):
            c = ws.cell(row=start_row, column=col_idx, value=header)
            c.fill = PatternFill("solid", fgColor=_argb(HEADER_DARK))
            c.font = Font(name="Calibri", size=10, bold=True, color=_argb(WHITE))
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        link_font = Font(name="Calibri", size=10, bold=True, color="FF0563C1", underline="single")
        body_font = Font(name="Calibri", size=10)
        alt_fill = PatternFill("solid", fgColor=_argb(LIGHT_FILL))
        default_fill = PatternFill("solid", fgColor=_argb(WHITE))
        thin = Side(style="thin", color=_argb(BORDER_GRAY))
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row_idx, entry in enumerate(index_entries, start=start_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else default_fill
            values = [
                entry["sheet"],
                entry["type"],
                entry["rows"],
                entry["description"],
                "Go to sheet",
            ]
            for col_idx, value in enumerate(values, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=value)
                c.fill = fill
                c.font = link_font if col_idx in (1, 5) else body_font
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = border

            target = self._sheet_hyperlink(entry["sheet"])
            ws.cell(row=row_idx, column=1).hyperlink = target
            ws.cell(row=row_idx, column=5).hyperlink = target

        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        ws.auto_filter.ref = f"A{start_row}:E{start_row + len(index_entries)}"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 70
        ws.column_dimensions["E"].width = 18

        # Source file footer
        footer_row = start_row + len(index_entries) + 3
        ws.cell(footer_row, 1, "Source Files").font = Font(name="Calibri", size=11, bold=True, color=_argb(WHITE))
        ws.cell(footer_row, 1).fill = PatternFill("solid", fgColor=_argb(HEADER_GREEN))
        for offset, (label, filename) in enumerate(source_files.items(), start=1):
            ws.cell(footer_row + offset, 1, label).font = Font(name="Calibri", size=10, bold=True)
            ws.cell(footer_row + offset, 2, filename).font = Font(name="Calibri", size=10)

        logger.info("Exporter: Workbook_Index tab written.")
        return title

    def _index_entry(self, sheet: str, description: str, sheet_type: str, rows: Any) -> Dict[str, Any]:
        return {
            "sheet": sheet,
            "description": description,
            "type": sheet_type,
            "rows": rows,
        }

    def _sheet_hyperlink(self, sheet_name: str) -> str:
        escaped = sheet_name.replace("'", "''")
        return f"#'{escaped}'!A1"

    # ------------------------------------------------------------------
    # Data tabs
    # ------------------------------------------------------------------

    def _write_data_tab(
        self,
        wb: Workbook,
        tab_name: str,
        df: pd.DataFrame,
        used_tab_names: List[str],
    ) -> str:
        """Write a standard data tab with OUTPUT_COLUMNS schema and return actual sheet title."""
        safe = safe_excel_tab_name(tab_name, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        df_out = self._standardize_output_df(df)

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row in enumerate(df_out.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")

        apply_data_tab_formatting(ws, OUTPUT_COLUMNS)
        logger.info("Exporter: Tab '%s' written — %d rows.", safe, len(df_out))
        return safe

    def _standardize_output_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """Return a copy of df with all OUTPUT_COLUMNS present and in order."""
        df_copy = df.copy() if df is not None else pd.DataFrame()
        for col in OUTPUT_COLUMNS:
            if col not in df_copy.columns:
                df_copy[col] = ""
        return df_copy[OUTPUT_COLUMNS].fillna("")

    # ------------------------------------------------------------------
    # Summary tab
    # ------------------------------------------------------------------

    def _write_summary(
        self,
        wb: Workbook,
        metrics: Dict[str, Any],
        source_files: Dict[str, str],
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        used_tab_names: List[str],
    ) -> str:
        safe = safe_excel_tab_name(SUMMARY_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        contact_matches = self._to_int(metrics.get("contact_matches", 0))
        contact_misses = self._to_int(metrics.get("contact_misses", 0))
        total_distinct = self._to_int(metrics.get("total_distinct_students", 0))
        total_assigned = self._to_int(metrics.get("total_assigned", 0))
        total_unmatched = self._to_int(metrics.get("total_unmatched", 0))
        contact_coverage = self._percent(contact_matches, contact_matches + contact_misses)
        assigned_pct = self._percent(total_assigned, total_distinct)
        unmatched_pct = self._percent(total_unmatched, total_distinct)

        rows = [
            ("◆  PROCESSING SUMMARY  ◆", ""),
            ("", ""),
            ("Application", f"{APP_NAME} v{APP_VERSION}"),
            ("Processing Timestamp", metrics.get("processing_timestamp", "")),
            ("Execution Duration", f"{self._to_float(metrics.get('execution_duration', 0)):.2f} seconds"),
            ("", ""),
            ("◆  EXECUTIVE KPIS  ◆", ""),
            ("Distinct At-Risk Students", total_distinct),
            ("Assigned to Groups", f"{total_assigned} ({assigned_pct})"),
            ("Unmatched", f"{total_unmatched} ({unmatched_pct})"),
            ("Contact Coverage", contact_coverage),
            ("Excluded Previously Assigned", metrics.get("excluded_previously_assigned", 0)),
            ("", ""),
            ("◆  INPUT FILE METRICS  ◆", ""),
            (SUMMARY_LABELS.get("total_input_rows", "Total Input Rows"), metrics.get("total_input_rows", 0)),
            (SUMMARY_LABELS.get("total_at_risk_rows", "Total At-Risk Rows"), metrics.get("total_at_risk_rows", 0)),
            (
                SUMMARY_LABELS.get("duplicate_course_rows_removed", "Duplicate Course Rows Removed"),
                metrics.get("duplicate_course_rows_removed", metrics.get("duplicate_rows_removed", 0)),
            ),
            (SUMMARY_LABELS.get("total_distinct_students", "Distinct At-Risk Students"), total_distinct),
            ("", ""),
            ("◆  CONTACT MATCHING  ◆", ""),
            (SUMMARY_LABELS.get("contact_matches", "Contact Matches"), contact_matches),
            (SUMMARY_LABELS.get("contact_misses", "Contact Misses"), contact_misses),
            ("Contact Coverage %", contact_coverage),
            ("", ""),
            ("◆  GROUP ASSIGNMENT  ◆", ""),
        ]

        for tab_name in group_order:
            df = group_data.get(tab_name, pd.DataFrame())
            pct = self._percent(len(df), total_distinct)
            rows.append((f"  {tab_name}", f"{len(df)} ({pct})"))

        rows.extend([
            ("", ""),
            ("◆  UNMATCHED BUCKETS  ◆", ""),
            (
                SUMMARY_LABELS.get("total_risk_1_2", "Risk 1–2"),
                len(group_data.get(UNMATCHED_LOW_TAB, pd.DataFrame())),
            ),
            (
                SUMMARY_LABELS.get("total_risk_3_plus", "Risk 3+"),
                len(group_data.get(UNMATCHED_HIGH_TAB, pd.DataFrame())),
            ),
            (SUMMARY_LABELS.get("total_unmatched", "Total Unmatched"), f"{total_unmatched} ({unmatched_pct})"),
            ("", ""),
            ("◆  SOURCE FILES  ◆", ""),
        ])

        for label, fname in source_files.items():
            rows.append((label, fname))

        rows.append(("", ""))
        rows.append(("Generated By", APP_NAME))

        for r_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=value)

        apply_summary_formatting(ws)
        self._apply_summary_finishing_touches(ws)

        try:
            enhancer = SummaryEnhancer()
            enhancer.enhance(ws, metrics, group_data, group_order)
        except Exception as exc:
            logger.warning("Exporter: Could not add summary charts: %s", exc)

        logger.info("Exporter: Summary tab written.")
        return safe

    def _apply_summary_finishing_touches(self, ws: Worksheet) -> None:
        """Small visual upgrade for summary sheets without depending on external config."""
        ws.freeze_panes = "A3"
        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 38
        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            if label and str(label).startswith("◆"):
                ws.row_dimensions[row].height = 22
            elif label:
                ws.row_dimensions[row].height = 18

    # ------------------------------------------------------------------
    # Missing Contacts tab
    # ------------------------------------------------------------------

    def _write_missing_contacts(
        self,
        wb: Workbook,
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        used_tab_names: List[str],
    ) -> tuple[Optional[str], int]:
        """Write a tab listing students with no phone or email found."""
        all_tabs = list(group_order) + [UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB]
        frames = []

        for tab in all_tabs:
            df = group_data.get(tab, pd.DataFrame())
            if df.empty:
                continue
            if "Phone Number" not in df.columns or "Email" not in df.columns:
                continue

            phone_blank = df["Phone Number"].fillna("").astype(str).str.strip() == ""
            email_blank = df["Email"].fillna("").astype(str).str.strip() == ""
            missing = df[phone_blank & email_blank].copy()

            if missing.empty:
                continue

            for col in ["Student Name", "Student ID", "Matched Group"]:
                if col not in missing.columns:
                    missing[col] = ""
            frames.append(missing[["Student Name", "Student ID", "Matched Group"]])

        if not frames:
            return None, 0

        missing_df = pd.concat(frames, ignore_index=True).drop_duplicates("Student ID")

        safe = safe_excel_tab_name(MISSING_CONTACTS_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        cols = ["Student Name", "Student ID", "Matched Group"]
        for c_idx, col in enumerate(cols, start=1):
            cell = ws.cell(1, c_idx, col)
            cell.fill = make_header_fill(HEADER_ORANGE)
            cell.font = make_header_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        body_font = make_body_font()
        alt_fill = PatternFill("solid", fgColor=_argb(LIGHT_BLUE))
        default_fill = PatternFill("solid", fgColor=_argb(WHITE))

        for r_idx, (_, row) in enumerate(missing_df.iterrows(), start=2):
            fill = alt_fill if r_idx % 2 == 0 else default_fill
            for c_idx, col in enumerate(cols, start=1):
                cell = ws.cell(r_idx, c_idx, str(row.get(col, "")))
                cell.fill = fill
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.freeze_panes = ws.cell(2, 1)
        ws.auto_filter.ref = f"A1:C{max(ws.max_row, 1)}"
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 24

        logger.info("Exporter: Missing_Contacts tab — %d students.", len(missing_df))
        return safe, len(missing_df)

    # ------------------------------------------------------------------
    # QA Log tab
    # ------------------------------------------------------------------

    def _write_qa_log(self, wb: Workbook, qa_log: QALog, used_tab_names: List[str]) -> str:
        safe = safe_excel_tab_name(QA_LOG_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        entries = qa_log.entries()

        for col_idx, col_name in enumerate(QA_LOG_COLUMNS, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, entry in enumerate(entries, start=2):
            for col_idx, col_name in enumerate(QA_LOG_COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=entry.get(col_name, ""))

        apply_qa_formatting(ws)
        logger.info("Exporter: QA_Log tab written — %d entries.", len(entries))
        return safe

    # ------------------------------------------------------------------
    # Processing Manifest tab
    # ------------------------------------------------------------------

    def _write_manifest(
        self,
        wb: Workbook,
        metrics: Dict[str, Any],
        source_files: Dict[str, str],
        used_tab_names: List[str],
    ) -> str:
        safe = safe_excel_tab_name(MANIFEST_TAB, used_tab_names)
        used_tab_names.append(safe)
        ws = wb.create_sheet(title=safe)

        rows = [
            ("◆  PROCESSING MANIFEST  ◆", ""),
            ("", ""),
            ("Application", APP_NAME),
            ("Version", APP_VERSION),
            ("Python Version", sys.version.split()[0]),
            ("Platform", platform.platform()),
            ("Processing Timestamp", metrics.get("processing_timestamp", "")),
            ("Execution Duration (s)", f"{self._to_float(metrics.get('execution_duration', 0)):.2f}"),
            ("", ""),
            ("◆  SOURCE FILES  ◆", ""),
        ]

        for label, fname in source_files.items():
            rows.append((label, fname))

        rows.extend([
            ("", ""),
            ("◆  ROW COUNTS  ◆", ""),
            ("Total Input Rows", metrics.get("total_input_rows", 0)),
            ("Total At-Risk Rows", metrics.get("total_at_risk_rows", 0)),
            ("Duplicate Course Rows Removed", metrics.get("duplicate_course_rows_removed", metrics.get("duplicate_rows_removed", 0))),
            ("Distinct At-Risk Students", metrics.get("total_distinct_students", 0)),
            ("Assigned to Groups", metrics.get("total_assigned", 0)),
            ("Unmatched", metrics.get("total_unmatched", 0)),
            ("Contact Matches", metrics.get("contact_matches", 0)),
            ("Contact Misses", metrics.get("contact_misses", 0)),
            ("Excluded Previously Assigned", metrics.get("excluded_previously_assigned", 0)),
            ("", ""),
            ("◆  OUTPUT FILE  ◆", ""),
            ("Output File", metrics.get("output_filename", "")),
        ])

        for r_idx, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r_idx, column=1, value=label)
            ws.cell(row=r_idx, column=2, value=value)

        apply_manifest_formatting(ws)
        logger.info("Exporter: Processing_Manifest tab written.")
        return safe

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _percent(self, numerator: Any, denominator: Any) -> str:
        numerator = self._to_float(numerator)
        denominator = self._to_float(denominator)
        if denominator <= 0:
            return "0.0%"
        return f"{(numerator / denominator) * 100:.1f}%"

    def _to_int(self, value: Any) -> int:
        try:
            return int(float(value or 0))
        except (TypeError, ValueError):
            return 0

    def _to_float(self, value: Any) -> float:
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0
