"""
summary_enhancer.py — Enhances the Summary tab in output workbooks with
executive KPI blocks, charts, group rankings, and risk visuals.
"""

import io
import logging
from typing import Dict, Any, List

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

from utils.config import UNMATCHED_LOW_TAB, UNMATCHED_HIGH_TAB
from utils.excel_utils import _argb

logger = logging.getLogger("intervention_sorter")

COLOR_BG = "#F4F6FB"
COLOR_PRIMARY = "#1F3864"
COLOR_ACCENT = "#2F5496"
COLOR_GREEN = "#2E7D32"
COLOR_RED = "#C62828"
COLOR_AMBER = "#F57F17"
COLOR_TEAL = "#00695C"
COLOR_GRAY = "#ECEFF1"

CHART_W = 400
CHART_H = 260


class SummaryEnhancer:
    """Adds charts and visual summaries to the Summary worksheet."""

    def enhance(
        self,
        ws: Worksheet,
        metrics: Dict[str, Any],
        group_data: Dict[str, pd.DataFrame],
        group_order: List[str],
        start_row: int = None,
    ) -> None:

        if start_row is None:
            start_row = ws.max_row + 3

        self._write_kpis(ws, metrics, start_row)
        start_row += 8

        group_counts = {
            tab: len(group_data.get(tab, pd.DataFrame())) for tab in group_order
        }

        group_counts[UNMATCHED_LOW_TAB] = len(
            group_data.get(UNMATCHED_LOW_TAB, pd.DataFrame())
        )

        group_counts[UNMATCHED_HIGH_TAB] = len(
            group_data.get(UNMATCHED_HIGH_TAB, pd.DataFrame())
        )

        group_counts = {k: v for k, v in group_counts.items() if v > 0}

        if group_counts:
            img = self._make_group_bar(group_counts)
            ws.add_image(img, f"A{start_row}")

            self._write_top_groups(
                ws,
                group_counts,
                start_row,
                start_col=9,
            )

            start_row += 20

        matched = metrics.get("contact_matches", 0)
        missed = metrics.get("contact_misses", 0)

        if matched + missed > 0:
            img2 = self._make_contact_donut(matched, missed)
            ws.add_image(img2, f"A{start_row}")

        all_students = pd.concat(
            [df for df in group_data.values() if not df.empty],
            ignore_index=True,
        )

        if not all_students.empty and "Risk Course Count" in all_students.columns:
            img3 = self._make_risk_distribution(all_students)
            ws.add_image(img3, f"E{start_row}")

    def _write_kpis(
        self,
        ws: Worksheet,
        metrics: Dict[str, Any],
        row: int,
    ) -> None:

        title_fill = PatternFill(
            "solid",
            fgColor=_argb(COLOR_PRIMARY),
        )

        metric_fill = PatternFill(
            "solid",
            fgColor=_argb(COLOR_GRAY),
        )

        title_font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color=_argb("FFFFFF"),
        )

        label_font = Font(
            name="Calibri",
            size=10,
            bold=True,
            color=_argb(COLOR_PRIMARY),
        )

        value_font = Font(
            name="Calibri",
            size=14,
            bold=True,
        )

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        c = ws.cell(row=row, column=1)
        c.value = "Executive Summary Metrics"
        c.fill = title_fill
        c.font = title_font
        c.alignment = Alignment(horizontal="center")

        kpis = [
            ("Students", metrics.get("total_distinct_students", 0)),
            ("Contact Matches", metrics.get("contact_matches", 0)),
            ("Contact Misses", metrics.get("contact_misses", 0)),
            ("Duplicates Removed", metrics.get("duplicate_course_rows_removed", 0)),
            ("Previously Assigned", metrics.get("excluded_previously_assigned", 0)),
            ("At-Risk Rows", metrics.get("total_at_risk_rows", 0)),
        ]

        current_col = 1

        for label, value in kpis:

            ws.merge_cells(
                start_row=row + 1,
                start_column=current_col,
                end_row=row + 1,
                end_column=current_col + 1,
            )

            ws.merge_cells(
                start_row=row + 2,
                start_column=current_col,
                end_row=row + 3,
                end_column=current_col + 1,
            )

            label_cell = ws.cell(row=row + 1, column=current_col)
            label_cell.value = label
            label_cell.fill = metric_fill
            label_cell.font = label_font
            label_cell.alignment = Alignment(horizontal="center")

            value_cell = ws.cell(row=row + 2, column=current_col)
            value_cell.value = value
            value_cell.fill = metric_fill
            value_cell.font = value_font
            value_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            ws.column_dimensions[get_column_letter(current_col)].width = 18

            current_col += 2

    def _write_top_groups(
        self,
        ws: Worksheet,
        group_counts: Dict[str, int],
        row: int,
        start_col: int = 9,
    ) -> None:

        sorted_groups = sorted(
            group_counts.items(),
            key=lambda x: x[1],
            reverse=True,
        )[:10]

        header_fill = PatternFill(
            "solid",
            fgColor=_argb(COLOR_ACCENT),
        )

        header_font = Font(
            bold=True,
            color=_argb("FFFFFF"),
        )

        ws.cell(row=row, column=start_col).value = "Top Groups"
        ws.cell(row=row, column=start_col).fill = header_fill
        ws.cell(row=row, column=start_col).font = header_font

        ws.cell(row=row, column=start_col + 1).value = "Students"
        ws.cell(row=row, column=start_col + 1).fill = header_fill
        ws.cell(row=row, column=start_col + 1).font = header_font

        for idx, (group, count) in enumerate(sorted_groups, start=1):
            ws.cell(row=row + idx, column=start_col).value = group
            ws.cell(row=row + idx, column=start_col + 1).value = count

        ws.column_dimensions[get_column_letter(start_col)].width = 28
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 12

    def _make_group_bar(self, group_counts: Dict[str, int]) -> XLImage:

        labels = list(group_counts.keys())
        counts = list(group_counts.values())

        colors = []

        for label in labels:
            if label == UNMATCHED_HIGH_TAB:
                colors.append(COLOR_RED)
            elif label == UNMATCHED_LOW_TAB:
                colors.append(COLOR_AMBER)
            else:
                colors.append(COLOR_ACCENT)

        fig, ax = plt.subplots(
            figsize=(5.8, max(2.8, len(labels) * 0.42 + 1)),
            facecolor=COLOR_BG,
        )

        ax.set_facecolor(COLOR_BG)

        bars = ax.barh(
            labels,
            counts,
            color=colors,
            edgecolor="white",
        )

        for bar, count in zip(bars, counts):
            ax.text(
                bar.get_width() + 0.2,
                bar.get_y() + bar.get_height() / 2,
                str(count),
                va="center",
                fontsize=9,
                fontweight="bold",
                color=COLOR_PRIMARY,
            )

        ax.set_xlabel("Students", fontsize=9)

        ax.set_title(
            "Students by Group",
            fontsize=11,
            fontweight="bold",
            color=COLOR_PRIMARY,
        )

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        ax.invert_yaxis()

        plt.tight_layout()

        return self._to_img(fig, CHART_W, CHART_H)

    def _make_contact_donut(self, matched: int, missed: int) -> XLImage:

        total = matched + missed
        pct = matched / total * 100 if total else 0

        fig, ax = plt.subplots(
            figsize=(3.6, 2.9),
            facecolor=COLOR_BG,
        )

        ax.set_facecolor(COLOR_BG)

        sizes = [matched, missed] if missed > 0 else [matched, 0.001]

        ax.pie(
            sizes,
            colors=[COLOR_GREEN, "#E0E0E0"],
            startangle=90,
            wedgeprops=dict(
                width=0.45,
                edgecolor="white",
                linewidth=1.5,
            ),
        )

        ax.text(
            0,
            0.1,
            f"{pct:.0f}%",
            ha="center",
            va="center",
            fontsize=16,
            fontweight="bold",
            color=COLOR_PRIMARY,
        )

        ax.text(
            0,
            -0.2,
            "with contact",
            ha="center",
            va="center",
            fontsize=8,
            color="#546E7A",
        )

        ax.set_title(
            "Contact Coverage",
            fontsize=11,
            fontweight="bold",
            color=COLOR_PRIMARY,
            pad=6,
        )

        plt.tight_layout()

        return self._to_img(fig, 300, 240)

    def _make_risk_distribution(
        self,
        df: pd.DataFrame,
    ) -> XLImage:

        counts = df["Risk Course Count"].value_counts().sort_index()

        labels = [f"{i} course{'s' if i != 1 else ''}" for i in counts.index]

        values = counts.values.tolist()

        colors = [COLOR_AMBER if i < 3 else COLOR_RED for i in counts.index]

        fig, ax = plt.subplots(
            figsize=(5.2, 2.9),
            facecolor=COLOR_BG,
        )

        ax.set_facecolor(COLOR_BG)

        bars = ax.bar(
            labels,
            values,
            color=colors,
            edgecolor="white",
        )

        for bar, val in zip(bars, values):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.5,
                str(val),
                ha="center",
                va="bottom",
                fontsize=9,
                fontweight="bold",
                color=COLOR_PRIMARY,
            )

        ax.set_ylabel("Students", fontsize=9)

        ax.set_title(
            "Risk Course Count Distribution",
            fontsize=11,
            fontweight="bold",
            color=COLOR_PRIMARY,
        )

        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

        plt.tight_layout()

        return self._to_img(fig, CHART_W, 220)

    def _to_img(
        self,
        fig,
        width: int,
        height: int,
    ) -> XLImage:

        buf = io.BytesIO()

        fig.savefig(
            buf,
            format="png",
            dpi=96,
            bbox_inches="tight",
            facecolor=COLOR_BG,
        )

        buf.seek(0)

        plt.close(fig)

        img = XLImage(buf)
        img.width = width
        img.height = height

        return img
