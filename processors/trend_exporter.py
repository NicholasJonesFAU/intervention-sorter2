"""
trend_exporter.py — Generates the Campaign Trend Report Excel workbook.

Tabs:
    Overview       — key stats + donut charts side by side
    Flow           — transition table + stacked bar chart
    Trajectories   — student state breakdown + horizontal bar chart
    By_Group       — group counts across checkpoints + grouped bar chart
"""

import io
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

from processors.trend_analyzer import TrendAnalyzer
from utils.config import (
    TRAJECTORY_LABELS,
    TRAJECTORY_COLORS,
    TREND_OUTPUT_FILENAME_PATTERN,
)
from utils.excel_utils import _argb

logger = logging.getLogger("intervention_sorter")

HEADER_DARK = "1F3864"
HEADER_FLOW = "2F5496"
HEADER_TRAJ = "375623"
HEADER_GROUP = "4A235A"
COLOR_PR1 = "#1565C0"
COLOR_MID = "#6A1B9A"
COLOR_PR2 = "#C62828"
COLOR_BG = "#F4F6FB"

# Standard chart dimensions in Excel pixels
CHART_W = 420
CHART_H = 260
DONUT_W = 220
DONUT_H = 200


class TrendExporter:

    def export(
        self, analyzer, output_path, pr1_name="PR1", mid_name="Midterm", pr2_name="PR2"
    ):
        logger.info("TrendExporter: Building workbook → '%s'", output_path.name)
        wb = Workbook()
        wb.remove(wb.active)

        overall = analyzer.overall_stats()
        flow_df = analyzer.checkpoint_flow()
        traj_df = analyzer.trajectory_breakdown()
        group_df = analyzer.group_breakdown()

        self._write_overview(wb, overall, pr1_name, mid_name, pr2_name)
        self._write_flow(wb, flow_df)
        self._write_trajectories(wb, traj_df)
        self._write_groups(wb, group_df)

        try:
            wb.save(output_path)
            logger.info("TrendExporter: Saved → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(
                f"Cannot save — file may be open in Excel.\n{exc}"
            ) from exc

    # ------------------------------------------------------------------
    # Overview tab — stats on left, three donuts side by side on right
    # ------------------------------------------------------------------

    def _write_overview(self, wb, overall, pr1_name, mid_name, pr2_name):
        ws = wb.create_sheet("Overview")

        # Banner row
        ws.merge_cells("A1:I1")
        c = ws["A1"]
        c.value = "Campaign Cycle — At-Risk Trend Report"
        c.font = Font(name="Calibri", size=13, bold=True, color=_argb("FFFFFF"))
        c.fill = PatternFill("solid", fgColor=_argb(HEADER_DARK))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Stats table in columns A-B
        body = Font(name="Calibri", size=10)
        bold = Font(name="Calibri", size=10, bold=True)
        stats = [
            ("Generated", datetime.now().strftime("%m/%d/%Y %I:%M %p")),
            ("", ""),
            ("Total Unique At-Risk Students", overall["total_unique_students"]),
            ("", ""),
            (f"{pr1_name} — Students at Risk", overall["pr1_count"]),
            (f"{mid_name} — Students at Risk", overall["mid_count"]),
            (f"{pr2_name} — Students at Risk", overall["pr2_count"]),
            ("", ""),
            (f"Recovered {pr1_name} → {mid_name}", overall["pr1_to_mid_recovered"]),
            (f"New at {mid_name}", overall["new_at_midterm"]),
            (f"Recovered {mid_name} → {pr2_name}", overall["mid_to_pr2_recovered"]),
            (f"New at {pr2_name}", overall["new_at_pr2"]),
        ]
        for r, (label, val) in enumerate(stats, start=2):
            lc = ws.cell(r, 1, label)
            lc.font = (
                bold
                if "Students at Risk" in str(label) or "Total" in str(label)
                else body
            )
            ws.cell(r, 2, val).font = body

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 10

        # Three donuts side by side — columns D, G, J
        total = overall["total_unique_students"] or 1
        checkpoints = [
            (pr1_name, overall["pr1_count"], COLOR_PR1, "D2"),
            (mid_name, overall["mid_count"], COLOR_MID, "G2"),
            (pr2_name, overall["pr2_count"], COLOR_PR2, "J2"),
        ]
        for name, count, color, anchor in checkpoints:
            img = self._make_donut(name, count, total, color, DONUT_W, DONUT_H)
            ws.add_image(img, anchor)

    # ------------------------------------------------------------------
    # Flow tab
    # ------------------------------------------------------------------

    def _write_flow(self, wb, flow_df):
        ws = wb.create_sheet("Checkpoint_Flow")
        self._write_table(ws, flow_df, HEADER_FLOW, "Student Flow Between Checkpoints")

        transitions = flow_df["Transition"].tolist()
        carried = flow_df["Carried Forward"].tolist()
        recovered = flow_df["Recovered"].tolist()
        new_vals = flow_df["New"].tolist()
        x = np.arange(len(transitions))

        fig, ax = plt.subplots(figsize=(5, 3), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        ax.bar(x, carried, label="Carried Forward", color=COLOR_PR1, alpha=0.85)
        ax.bar(
            x, recovered, bottom=carried, label="Recovered", color="#2E7D32", alpha=0.85
        )
        ax.bar(
            x,
            new_vals,
            bottom=[c + r for c, r in zip(carried, recovered)],
            label="New",
            color=COLOR_MID,
            alpha=0.85,
        )
        ax.set_xticks(x)
        ax.set_xticklabels(transitions, fontsize=9)
        ax.set_ylabel("Students", fontsize=9)
        ax.set_title("Student Flow", fontsize=11, fontweight="bold", color="#1F3864")
        ax.legend(fontsize=8, framealpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()
        img = self._fig_to_img(fig, CHART_W, CHART_H)
        ws.add_image(img, f"A{len(flow_df) + 5}")

    # ------------------------------------------------------------------
    # Trajectories tab
    # ------------------------------------------------------------------

    def _write_trajectories(self, wb, traj_df):
        ws = wb.create_sheet("Trajectories")
        self._write_table(ws, traj_df, HEADER_TRAJ, "Student Trajectory Breakdown")

        df_sorted = traj_df.sort_values("Count", ascending=True)
        labels = df_sorted["Trajectory"].tolist()
        counts = df_sorted["Count"].tolist()
        key_map = {v: k for k, v in TRAJECTORY_LABELS.items()}
        colors = [TRAJECTORY_COLORS.get(key_map.get(l, ""), "#90CAF9") for l in labels]

        fig, ax = plt.subplots(
            figsize=(6, max(2.5, len(labels) * 0.35 + 0.5)), facecolor=COLOR_BG
        )
        ax.set_facecolor(COLOR_BG)
        bars = ax.barh(labels, counts, color=colors, edgecolor="white")
        for bar, count in zip(bars, counts):
            if count > 0:
                ax.text(
                    bar.get_width() + 0.2,
                    bar.get_y() + bar.get_height() / 2,
                    str(count),
                    va="center",
                    fontsize=8,
                    fontweight="bold",
                    color="#1F3864",
                )
        ax.set_xlabel("Students", fontsize=9)
        ax.set_title(
            "Student Trajectories", fontsize=11, fontweight="bold", color="#1F3864"
        )
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()
        ws.add_image(
            self._fig_to_img(fig, CHART_W, max(CHART_H, len(labels) * 28 + 60)),
            f"A{len(traj_df) + 5}",
        )

    # ------------------------------------------------------------------
    # Group tab
    # ------------------------------------------------------------------

    def _write_groups(self, wb, group_df):
        ws = wb.create_sheet("By_Group")
        if group_df.empty:
            ws["A1"] = "No group data available."
            return
        self._write_table(ws, group_df, HEADER_GROUP, "At-Risk Counts by Group")

        groups = group_df["Group"].tolist()
        pr1_n = group_df["PR1 Count"].tolist()
        mid_n = group_df["Midterm Count"].tolist()
        pr2_n = group_df["PR2 Count"].tolist()
        x = np.arange(len(groups))
        w = 0.25

        fig, ax = plt.subplots(
            figsize=(max(5, len(groups) * 0.8 + 1), 3.5), facecolor=COLOR_BG
        )
        ax.set_facecolor(COLOR_BG)
        ax.bar(x - w, pr1_n, w, label="PR1", color=COLOR_PR1, alpha=0.85)
        ax.bar(x, mid_n, w, label="Midterm", color=COLOR_MID, alpha=0.85)
        ax.bar(x + w, pr2_n, w, label="PR2", color=COLOR_PR2, alpha=0.85)
        ax.set_xticks(x)
        ax.set_xticklabels(groups, rotation=25, ha="right", fontsize=9)
        ax.set_ylabel("Students", fontsize=9)
        ax.set_title(
            "At-Risk by Group", fontsize=11, fontweight="bold", color="#1F3864"
        )
        ax.legend(fontsize=8, framealpha=0.5)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()
        ws.add_image(self._fig_to_img(fig, CHART_W, CHART_H), f"A{len(group_df) + 5}")

    # ------------------------------------------------------------------
    # Shared table writer
    # ------------------------------------------------------------------

    def _write_table(self, ws, df, header_color, title):
        ws.cell(1, 1, title)
        ws.cell(1, 1).font = Font(
            name="Calibri", size=11, bold=True, color=_argb("FFFFFF")
        )
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=_argb(header_color))
        ws.row_dimensions[1].height = 22

        for c_idx, col in enumerate(df.columns, 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_argb("FFFFFF"))
            cell.fill = PatternFill("solid", fgColor=_argb(header_color))
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = ws.cell(3, 1)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}2"

        alt = PatternFill("solid", fgColor=_argb("DCE6F1"))
        body = Font(name="Calibri", size=10)
        for r_idx, row in enumerate(df.itertuples(index=False), start=3):
            fill = (
                alt if r_idx % 2 == 1 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            )
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, str(val) if pd.notna(val) else "")
                cell.fill = fill
                cell.font = body
                cell.alignment = Alignment(
                    horizontal="center" if c_idx > 1 else "left", vertical="top"
                )

        for c_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(col_name), 8)
            for r_idx in range(3, min(ws.max_row + 1, 200)):
                val = ws.cell(r_idx, c_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 45)

    # ------------------------------------------------------------------
    # Chart helpers
    # ------------------------------------------------------------------

    def _make_donut(self, title, count, total, color, width, height):
        pct = count / total * 100 if total else 0
        other = max(total - count, 0)
        fig, ax = plt.subplots(figsize=(2.2, 1.9), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        sizes = [count, other] if other > 0 else [count, 0.001]
        ax.pie(
            sizes,
            colors=[color, "#E0E0E0"],
            startangle=90,
            wedgeprops=dict(width=0.45, edgecolor="white", linewidth=1),
        )
        ax.text(
            0,
            0.08,
            str(count),
            ha="center",
            va="center",
            fontsize=14,
            fontweight="bold",
            color="#1F3864",
        )
        ax.text(
            0,
            -0.22,
            f"{pct:.1f}%",
            ha="center",
            va="center",
            fontsize=8,
            color="#546E7A",
        )
        ax.set_title(title, fontsize=9, fontweight="bold", color="#1F3864", pad=4)
        return self._fig_to_img(fig, width, height)

    def _fig_to_img(self, fig, width, height) -> XLImage:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=96, bbox_inches="tight", facecolor=COLOR_BG)
        buf.seek(0)
        plt.close(fig)
        img = XLImage(buf)
        img.width = width
        img.height = height
        return img
