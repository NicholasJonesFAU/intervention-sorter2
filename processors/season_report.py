"""
season_report.py — Generates the end-of-semester Master Season Report.

Reads PR1, Midterm, and PR2 output workbooks and produces a single
comprehensive report with:
  - Season Summary tab — key stats, recovery rates, overall counts
  - Master Student List — every unique student, one row, all checkpoints
  - By Group — group-level breakdown across all checkpoints
  - Visuals — recovery funnel, persistence by group, contact coverage
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Set, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

from processors.trend_analyzer import TrendAnalyzer
from utils.config import SUMMARY_TAB, QA_LOG_TAB, MANIFEST_TAB
from utils.excel_utils import _argb

logger = logging.getLogger("intervention_sorter")

SKIP_TABS = {SUMMARY_TAB, QA_LOG_TAB, MANIFEST_TAB, "QA_Log", "Processing_Manifest"}
COLOR_BG = "#F4F6FB"
COLOR_DARK = "1F3864"
COLOR_MID_H = "2F5496"
COLOR_GREEN = "375623"
COLOR_PR1 = "#1565C0"
COLOR_MID = "#6A1B9A"
COLOR_PR2 = "#C62828"
CHART_W, CHART_H = 420, 260


class SeasonReportGenerator:

    def generate(
        self,
        pr1_path: Optional[Path],
        mid_path: Optional[Path],
        pr2_path: Optional[Path],
        output_path: Path,
        season_name: str = "",
        pr1_label: str = "Progress Report 1",
        mid_label: str = "Midterm",
        pr2_label: str = "Progress Report 2",
    ) -> None:
        logger.info("SeasonReport: Generating → '%s'", output_path.name)

        # Load data from workbooks
        pr1_df = (
            self._read_workbook(pr1_path, pr1_label) if pr1_path else pd.DataFrame()
        )
        mid_df = (
            self._read_workbook(mid_path, mid_label) if mid_path else pd.DataFrame()
        )
        pr2_df = (
            self._read_workbook(pr2_path, pr2_label) if pr2_path else pd.DataFrame()
        )

        # Use trend analyzer for trajectory stats
        analyzer = TrendAnalyzer()
        analyzer.load(pr1_path, mid_path, pr2_path)
        overall = analyzer.overall_stats()
        traj_df = analyzer.trajectory_breakdown()
        group_df = analyzer.group_breakdown()
        flow_df = analyzer.checkpoint_flow()

        # Build master student list
        master_df = self._build_master_list(
            pr1_df, mid_df, pr2_df, pr1_label, mid_label, pr2_label
        )

        wb = Workbook()
        wb.remove(wb.active)

        self._write_season_summary(
            wb, overall, traj_df, flow_df, season_name, pr1_label, mid_label, pr2_label
        )
        self._write_master_list(wb, master_df, pr1_label, mid_label, pr2_label)
        self._write_by_group(wb, group_df)

        try:
            wb.save(output_path)
            logger.info("SeasonReport: Saved → '%s'", output_path)
        except PermissionError as exc:
            raise RuntimeError(f"Cannot save — file may be open: {exc}") from exc

    # ------------------------------------------------------------------
    # Season Summary tab
    # ------------------------------------------------------------------

    def _write_season_summary(
        self,
        wb,
        overall,
        traj_df,
        flow_df,
        season_name,
        pr1_label,
        mid_label,
        pr2_label,
    ):
        ws = wb.create_sheet("Season_Summary")

        # Banner
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value = (
            f"End-of-Semester Report — {season_name}"
            if season_name
            else "End-of-Semester Report"
        )
        c.font = Font(name="Calibri", size=14, bold=True, color=_argb("FFFFFF"))
        c.fill = PatternFill("solid", fgColor=_argb(COLOR_DARK))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        body = Font(name="Calibri", size=10)
        bold = Font(name="Calibri", size=10, bold=True)
        head = Font(name="Calibri", size=11, bold=True, color=_argb("FFFFFF"))

        def section(row, title, color):
            ws.cell(row, 1, title).font = head
            ws.cell(row, 1).fill = PatternFill("solid", fgColor=_argb(color))
            ws.row_dimensions[row].height = 20

        def stat(row, label, value):
            ws.cell(row, 1, label).font = bold
            ws.cell(row, 2, value).font = body

        r = 2
        section(r, "◆  Checkpoint Overview", COLOR_MID_H)
        r += 1
        stat(r, f"{pr1_label} — At-Risk Students", overall["pr1_count"])
        r += 1
        stat(r, f"{mid_label} — At-Risk Students", overall["mid_count"])
        r += 1
        stat(r, f"{pr2_label} — At-Risk Students", overall["pr2_count"])
        r += 1
        stat(
            r,
            "Total Unique Students (any checkpoint)",
            overall["total_unique_students"],
        )
        r += 2

        section(r, "◆  Recovery", COLOR_GREEN)
        r += 1
        stat(r, f"Recovered {pr1_label} → {mid_label}", overall["pr1_to_mid_recovered"])
        r += 1
        stat(r, f"New at {mid_label}", overall["new_at_midterm"])
        r += 1
        stat(r, f"Recovered {mid_label} → {pr2_label}", overall["mid_to_pr2_recovered"])
        r += 1
        stat(r, f"New at {pr2_label}", overall["new_at_pr2"])
        r += 2

        section(r, "◆  Trajectories", "843C0C")
        r += 1
        for _, row_data in traj_df.iterrows():
            stat(r, row_data["Trajectory"], f"{row_data['Count']}  ({row_data['Pct']})")
            r += 1
        r += 1

        stat(r, "Generated", datetime.now().strftime("%m/%d/%Y %I:%M %p"))

        ws.column_dimensions["A"].width = 42
        ws.column_dimensions["B"].width = 22

        # Add recovery funnel chart
        img = self._make_funnel(
            [overall["pr1_count"], overall["mid_count"], overall["pr2_count"]],
            [pr1_label, mid_label, pr2_label],
        )
        ws.add_image(img, "D2")

        # Trajectory chart
        img2 = self._make_trajectory_chart(traj_df)
        ws.add_image(img2, "D18")

    # ------------------------------------------------------------------
    # Master Student List tab
    # ------------------------------------------------------------------

    def _write_master_list(self, wb, master_df, pr1_label, mid_label, pr2_label):
        ws = wb.create_sheet("Master_Student_List")

        if master_df.empty:
            ws["A1"] = "No student data available."
            return

        # Header
        ws.cell(1, 1, "Master At-Risk Student List — All Checkpoints")
        ws.cell(1, 1).font = Font(
            name="Calibri", size=12, bold=True, color=_argb("FFFFFF")
        )
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=_argb(COLOR_DARK))
        ws.row_dimensions[1].height = 22

        cols = list(master_df.columns)
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_argb("FFFFFF"))
            cell.fill = PatternFill("solid", fgColor=_argb(COLOR_DARK))
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = ws.cell(3, 1)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}2"

        alt = PatternFill("solid", fgColor=_argb("DCE6F1"))
        body = Font(name="Calibri", size=10)
        for r_idx, row in enumerate(master_df.itertuples(index=False), start=3):
            fill = (
                alt if r_idx % 2 == 1 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            )
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, str(val) if pd.notna(val) else "")
                cell.fill = fill
                cell.font = body
                cell.alignment = Alignment(
                    horizontal="center" if c_idx > 2 else "left", vertical="top"
                )

        for c_idx, col in enumerate(cols, 1):
            max_len = max(len(col), 8)
            for r_idx in range(3, min(ws.max_row + 1, 300)):
                val = ws.cell(r_idx, c_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 3, 40)

    # ------------------------------------------------------------------
    # By Group tab
    # ------------------------------------------------------------------

    def _write_by_group(self, wb, group_df):
        ws = wb.create_sheet("By_Group")

        if group_df.empty:
            ws["A1"] = "No group data available."
            return

        ws.cell(1, 1, "At-Risk Counts by Group — Full Semester")
        ws.cell(1, 1).font = Font(
            name="Calibri", size=12, bold=True, color=_argb("FFFFFF")
        )
        ws.cell(1, 1).fill = PatternFill("solid", fgColor=_argb("4A235A"))
        ws.row_dimensions[1].height = 22

        cols = list(group_df.columns)
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(name="Calibri", size=10, bold=True, color=_argb("FFFFFF"))
            cell.fill = PatternFill("solid", fgColor=_argb("4A235A"))
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = ws.cell(3, 1)
        alt = PatternFill("solid", fgColor=_argb("DCE6F1"))
        body = Font(name="Calibri", size=10)
        for r_idx, row in enumerate(group_df.itertuples(index=False), start=3):
            fill = (
                alt if r_idx % 2 == 1 else PatternFill("solid", fgColor=_argb("FFFFFF"))
            )
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, str(val) if pd.notna(val) else "")
                cell.fill = fill
                cell.font = body
                cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

        for c_idx, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = min(len(col) + 6, 30)

        # Grouped bar chart
        if not group_df.empty and all(
            c in group_df.columns for c in ["PR1 Count", "Midterm Count", "PR2 Count"]
        ):
            groups = group_df["Group"].tolist()
            x = np.arange(len(groups))
            w = 0.25
            fig, ax = plt.subplots(
                figsize=(max(5, len(groups) * 0.8 + 1), 3.5), facecolor=COLOR_BG
            )
            ax.set_facecolor(COLOR_BG)
            ax.bar(
                x - w,
                group_df["PR1 Count"].tolist(),
                w,
                label="PR1",
                color=COLOR_PR1,
                alpha=0.85,
            )
            ax.bar(
                x,
                group_df["Midterm Count"].tolist(),
                w,
                label="Mid",
                color=COLOR_MID,
                alpha=0.85,
            )
            ax.bar(
                x + w,
                group_df["PR2 Count"].tolist(),
                w,
                label="PR2",
                color=COLOR_PR2,
                alpha=0.85,
            )
            ax.set_xticks(x)
            ax.set_xticklabels(groups, rotation=25, ha="right", fontsize=9)
            ax.set_ylabel("Students", fontsize=9)
            ax.set_title(
                "At-Risk by Group — Full Semester",
                fontsize=11,
                fontweight="bold",
                color="#1F3864",
            )
            ax.legend(fontsize=8, framealpha=0.5)
            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            plt.tight_layout()
            img = self._to_img(fig, CHART_W, CHART_H)
            ws.add_image(img, f"A{len(group_df) + 5}")

    # ------------------------------------------------------------------
    # Data helpers
    # ------------------------------------------------------------------

    def _read_workbook(self, path: Path, label: str) -> pd.DataFrame:
        """Read all student rows from a generated output workbook."""
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
        except Exception as exc:
            logger.warning("SeasonReport: Cannot open %s: %s", label, exc)
            return pd.DataFrame()

        frames = []
        for sheet in xl.sheet_names:
            if sheet in SKIP_TABS:
                continue
            try:
                df = xl.parse(sheet, dtype=str)
                df.columns = [str(c).strip() for c in df.columns]
                if "Student ID" not in df.columns:
                    continue
                df["_checkpoint"] = label
                df["_group"] = sheet
                frames.append(df)
            except Exception:
                pass

        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def _build_master_list(
        self, pr1_df, mid_df, pr2_df, pr1_label, mid_label, pr2_label
    ) -> pd.DataFrame:
        """
        One row per unique student showing which checkpoints they appeared in,
        their group, name, contact info.
        """
        all_frames = []
        for df, label in [
            (pr1_df, pr1_label),
            (mid_df, mid_label),
            (pr2_df, pr2_label),
        ]:
            if not df.empty:
                all_frames.append(
                    df[
                        [
                            "Student ID",
                            "_checkpoint",
                            "_group",
                            "Student Name",
                            "Phone Number",
                            "Email",
                            "Matched Group",
                        ]
                    ].copy()
                    if all(
                        c in df.columns
                        for c in [
                            "Student Name",
                            "Phone Number",
                            "Email",
                            "Matched Group",
                        ]
                    )
                    else df[["Student ID", "_checkpoint", "_group"]].copy()
                )

        if not all_frames:
            return pd.DataFrame()

        combined = pd.concat(all_frames, ignore_index=True)

        # Pivot to get one row per student
        pivot = (
            combined.groupby("Student ID")
            .agg(
                Student_Name=(
                    ("Student Name", "first")
                    if "Student Name" in combined.columns
                    else ("Student ID", "first")
                ),
                Phone=(
                    ("Phone Number", "first")
                    if "Phone Number" in combined.columns
                    else ("Student ID", "first")
                ),
                Email=(
                    ("Email", "first")
                    if "Email" in combined.columns
                    else ("Student ID", "first")
                ),
            )
            .reset_index()
            if all(
                c in combined.columns for c in ["Student Name", "Phone Number", "Email"]
            )
            else combined[["Student ID"]].drop_duplicates()
        )

        # Add checkpoint flags
        for label, df in [
            (pr1_label, pr1_df),
            (mid_label, mid_df),
            (pr2_label, pr2_df),
        ]:
            if not df.empty and "Student ID" in df.columns:
                ids_in = set(
                    df["Student ID"].dropna().astype(str).str.strip().str.upper()
                )
                col_name = label[:15]  # Truncate for readability
                pivot[col_name] = pivot["Student ID"].apply(
                    lambda x: "✓" if str(x).strip().upper() in ids_in else ""
                )

        # Count appearances
        checkpoint_cols = [
            c
            for c in pivot.columns
            if c not in ["Student ID", "Student_Name", "Phone", "Email"]
        ]
        pivot["Appearances"] = pivot[checkpoint_cols].apply(
            lambda row: sum(1 for v in row if v == "✓"), axis=1
        )

        # Sort by appearances desc then name
        sort_cols = ["Appearances"]
        if "Student_Name" in pivot.columns:
            sort_cols.append("Student_Name")
        pivot = pivot.sort_values(
            sort_cols, ascending=[False] + [True] * (len(sort_cols) - 1)
        )

        # Clean up column names
        pivot = pivot.rename(
            columns={
                "Student_Name": "Student Name",
                "Phone": "Phone Number",
            }
        )

        return pivot.reset_index(drop=True)

    # ------------------------------------------------------------------
    # Chart helpers
    # ------------------------------------------------------------------

    def _make_funnel(self, counts, labels) -> XLImage:
        """Simple horizontal funnel showing count at each checkpoint."""
        colors = [COLOR_PR1, COLOR_MID, COLOR_PR2]
        fig, ax = plt.subplots(figsize=(5, 2.8), facecolor=COLOR_BG)
        ax.set_facecolor(COLOR_BG)
        y = np.arange(len(labels))
        bars = ax.barh(y, counts, color=colors, edgecolor="white", height=0.5)
        for bar, count, label in zip(bars, counts, labels):
            ax.text(
                bar.get_width() + max(counts) * 0.02,
                bar.get_y() + bar.get_height() / 2,
                f"{count:,}",
                va="center",
                fontsize=10,
                fontweight="bold",
                color="#1F3864",
            )
        ax.set_yticks(y)
        ax.set_yticklabels(labels, fontsize=10)
        ax.set_xlabel("Students at Risk", fontsize=9)
        ax.set_title(
            "At-Risk Population by Checkpoint",
            fontsize=11,
            fontweight="bold",
            color="#1F3864",
        )
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.invert_yaxis()
        plt.tight_layout()
        return self._to_img(fig, CHART_W, 200)

    def _make_trajectory_chart(self, traj_df) -> XLImage:
        from utils.config import TRAJECTORY_LABELS, TRAJECTORY_COLORS

        df = traj_df.sort_values("Count", ascending=True)
        labels = df["Trajectory"].tolist()
        counts = df["Count"].tolist()
        key_map = {v: k for k, v in TRAJECTORY_LABELS.items()}
        colors = [TRAJECTORY_COLORS.get(key_map.get(l, ""), "#90CAF9") for l in labels]

        fig, ax = plt.subplots(
            figsize=(5.5, max(2.5, len(labels) * 0.38 + 0.5)), facecolor=COLOR_BG
        )
        ax.set_facecolor(COLOR_BG)
        bars = ax.barh(labels, counts, color=colors, edgecolor="white")
        for bar, count in zip(bars, counts):
            if count > 0:
                ax.text(
                    bar.get_width() + 0.2,
                    bar.get_y() + bar.get_height() / 2,
                    str(count),
                    va="center",
                    fontsize=8,
                    fontweight="bold",
                    color="#1F3864",
                )
        ax.set_xlabel("Students", fontsize=9)
        ax.set_title(
            "Student Trajectories", fontsize=11, fontweight="bold", color="#1F3864"
        )
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        plt.tight_layout()
        return self._to_img(fig, CHART_W, max(CHART_H, len(labels) * 28 + 60))

    def _to_img(self, fig, width, height) -> XLImage:
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=96, bbox_inches="tight", facecolor=COLOR_BG)
        buf.seek(0)
        plt.close(fig)
        img = XLImage(buf)
        img.width = width
        img.height = height
        return img
